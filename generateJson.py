from openpyxl import load_workbook
import os
from datetime import datetime
import requests
import json
import re
import uuid
import pickle
import os.path
import string
allowed=["dictionary"]

def main():
	xlFile=load_workbook("dictionary.xlsx")
	
	index=0
	#fileURL="https://drive.google.com/file/d/1GPWteQTcT5xD-Riild6OT0o6BSwMlS3f/view?usp=sharing"
	#getPublicURL(fileURL, google_drive_service)
	for sheetname in xlFile.sheetnames:	
		print(sheetname)	
		if(sheetname in allowed):
			generateJSON(xlFile, sheetname)
				
	
	#xlFile.save("Championship Questions V2.xlsx")

def formatString(keyname):
	retval=""
	for newkey in keyname:
		if(newkey.isalpha()):
				retval=retval+newkey
	return retval	


def generateJSON(fileref, shtname):
	sheet = fileref[shtname]

	
	rowStart=2
	rowEnd=54553
	#rows=122
	offset=1
	myDict={}
	#firstWord=""
	for row in range(rowStart,rowEnd):
		print(row)
		word=sheet.cell(row, offset+0).value
		word=word.lower()
		word=formatString(word)
		definition= sheet.cell(row, offset+2).value  
		blankObject={} 
		if (word in myDict): 
			blankObject=myDict[word]
			#blankObject["definitions"]=[]
			blankObject["definitions"].append(definition)
			myDict[word] = blankObject
		else: 
			blankObject["definitions"]=[]
			blankObject["definitions"].append(definition)
			myDict[word] = blankObject

	for key, value in myDict.items():
		filename=key+".json"
		with open(filename, "w") as jsonfile:
			jsonfile.write(json.dumps(value))
	
	
		


main()